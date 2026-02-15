from __future__ import annotations

import csv
import importlib.util
import re
import zlib
from datetime import datetime
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

RE_CNPJ = re.compile(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}")
RE_DATA_MM_AAAA = re.compile(r"\b\d{2}/\d{4}\b")
RE_PERIODO = re.compile(r"\b(?:\d{2}/\d{4}|[1-4][º°]\s*Trimestre/\d{4})\b", re.IGNORECASE)


@dataclass
class BlocoDebitoCredito:
    codigo_receita: str = ""
    descricao: str = ""
    periodo_apuracao: str = ""
    cnpj_debito: str = ""
    municipio_debito: str = ""
    debito_apurado: str = ""
    creditos_compensacao: str = ""
    saldo_a_pagar: str = ""
    deducoes: List[str] = field(default_factory=list)
    outros_campos: List[str] = field(default_factory=list)


@dataclass
class Compensacao:
    codigo_receita: str = ""
    descricao_receita: str = ""
    periodo_apuracao: str = ""
    numero_processo: str = ""
    tipo: str = ""
    valor: str = ""


def _decode_pdf_text_token(hex_text: Optional[bytes], literal_text: Optional[bytes]) -> str:
    if hex_text:
        return bytes.fromhex(hex_text.decode("ascii")).decode("latin-1", errors="ignore")

    raw = literal_text or b""
    out = bytearray()
    idx = 0

    while idx < len(raw):
        cur = raw[idx]
        if cur == 92 and idx + 1 < len(raw):
            nxt = raw[idx + 1]
            escapes = {
                ord("n"): ord("\n"),
                ord("r"): ord("\r"),
                ord("t"): ord("\t"),
                ord("b"): 8,
                ord("f"): 12,
                ord("("): ord("("),
                ord(")"): ord(")"),
                ord("\\"): ord("\\"),
            }
            if nxt in escapes:
                out.append(escapes[nxt])
                idx += 2
                continue

        out.append(cur)
        idx += 1

    return out.decode("latin-1", errors="ignore")


def _extract_pdf_objects(pdf_bytes: bytes) -> Dict[int, bytes]:
    found: Dict[int, bytes] = {}
    for m in re.finditer(rb"(\d+)\s+0\s+obj(.*?)endobj", pdf_bytes, re.S):
        obj_num = int(m.group(1))
        found[obj_num] = m.group(2)
    return found


def _page_content_order(objs: Dict[int, bytes]) -> List[int]:
    order: List[int] = []
    for _, body in sorted(objs.items(), key=lambda item: item[0]):
        if b"/Type /Page" not in body:
            continue

        m = re.search(rb"/Contents\s+(\d+)\s+0\s+R", body)
        if m:
            order.append(int(m.group(1)))

    return order


def _inflate_stream(obj_body: bytes) -> Optional[bytes]:
    stream_start = re.search(rb"stream\r?\n", obj_body)
    stream_end = obj_body.rfind(b"endstream")
    if not stream_start or stream_end < 0:
        return None

    content = obj_body[stream_start.end() : stream_end]

    if b"/FlateDecode" in obj_body:
        try:
            content = zlib.decompress(content)
        except zlib.error:
            return None

    return content


def _extract_lines_from_stream(stream: bytes) -> List[str]:
    token_re = re.compile(
        rb"1\s+0\s+0\s+1\s+([\d\.-]+)\s+([\d\.-]+)\s+Tm\s*(?:<([0-9A-Fa-f]+)>|\((.*?)\))\s*Tj",
        re.S,
    )

    items: List[Tuple[float, float, str]] = []
    for m in token_re.finditer(stream):
        x = float(m.group(1))
        y = float(m.group(2))
        text = _decode_pdf_text_token(m.group(3), m.group(4)).strip()
        if text:
            items.append((x, y, text))

    items.sort(key=lambda value: (-value[1], value[0]))

    grouped: List[Dict[str, object]] = []
    for x, y, text in items:
        attached = False
        for line in grouped:
            if abs(float(line["y"]) - y) < 1.2:
                line["parts"].append((x, text))
                attached = True
                break

        if not attached:
            grouped.append({"y": y, "parts": [(x, text)]})

    lines: List[str] = []
    for line in grouped:
        parts = sorted(line["parts"], key=lambda value: value[0])
        lines.append(" ".join(t for _, t in parts).strip())

    return lines


def extract_pdf_lines(pdf_path: Path) -> List[str]:
    pdf_bytes = pdf_path.read_bytes()
    objects = _extract_pdf_objects(pdf_bytes)
    contents = _page_content_order(objects)

    lines: List[str] = []
    for content_obj in contents:
        body = objects.get(content_obj)
        if not body:
            continue

        stream = _inflate_stream(body)
        if not stream:
            continue

        lines.extend(_extract_lines_from_stream(stream))

    return lines


def extract_header(lines: List[str]) -> Dict[str, str]:
    full_text = "\n".join(lines)
    cnpj_match = RE_CNPJ.search(full_text)

    contributor_name = ""
    for idx, line in enumerate(lines):
        if "Nome do Contribuinte" not in line:
            continue

        name_parts = []

        prev_line = lines[idx - 1].strip() if idx > 0 else ""
        if prev_line and "RELATÓRIO" not in prev_line and "SECRETARIA" not in prev_line:
            name_parts.append(prev_line)

        line_part = line.split("Nome do Contribuinte", 1)[-1]
        line_part = line_part.split("CNPJ", 1)[0].strip()
        if line_part:
            name_parts.append(line_part)

        if idx + 1 < len(lines):
            next_line = lines[idx + 1].strip()
            if next_line and "Período" not in next_line and "Data/Hora" not in next_line:
                name_parts.append(next_line)

        candidate = re.sub(r"\s+", " ", " ".join(name_parts)).strip(" -")
        if candidate:
            contributor_name = candidate
            break

    return {
        "nome_contribuinte": contributor_name,
        "cnpj": cnpj_match.group(0) if cnpj_match else "",
    }


def extract_debits_credits_and_offsets(lines: List[str]) -> Tuple[List[BlocoDebitoCredito], List[Compensacao]]:
    blocks: List[BlocoDebitoCredito] = []
    offsets: List[Compensacao] = []

    current: Optional[BlocoDebitoCredito] = None
    waiting_period = False
    in_offsets = False

    for line in lines:
        if "Débito Apurado e Crédito Vinculado" in line:
            if current and current.codigo_receita:
                blocks.append(current)
            current = BlocoDebitoCredito()
            waiting_period = False
            in_offsets = False
            continue

        if not current:
            continue

        code_match = re.search(r"Código da Receita\s+([0-9]{4}-[0-9]{2})\s+Descrição\s+(.+)$", line)
        if code_match:
            current.codigo_receita = code_match.group(1).strip()
            current.descricao = code_match.group(2).strip()
            continue

        debit_entity_match = re.search(
            r"CNPJ Débito\s+(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})\s+Município Débito\s+(.+)$",
            line,
        )
        if debit_entity_match:
            current.cnpj_debito = debit_entity_match.group(1).strip()
            current.municipio_debito = debit_entity_match.group(2).strip()
            continue

        if "Período Apuração" in line:
            period_match = RE_PERIODO.search(line)
            if period_match:
                current.periodo_apuracao = period_match.group(0)
                waiting_period = False
            else:
                waiting_period = True
            continue

        if waiting_period and RE_PERIODO.search(line):
            current.periodo_apuracao = RE_PERIODO.search(line).group(0)
            waiting_period = False
            continue

        debit_match = re.search(r"^Débito Apurado\s+(.+)$", line)
        if debit_match:
            current.debito_apurado = debit_match.group(1).strip()
            continue

        offset_credit_match = re.search(r"^Créditos Compensação:\s+(.+)$", line)
        if offset_credit_match:
            current.creditos_compensacao = offset_credit_match.group(1).strip()
            continue

        balance_match = re.search(r"^Saldo a Pagar\s+(.+)$", line)
        if balance_match:
            current.saldo_a_pagar = balance_match.group(1).strip()
            continue

        deduction_match = re.search(r"^Deduções\s+(.+)$", line)
        if deduction_match:
            current.deducoes.append(deduction_match.group(1).strip())
            continue

        if line.strip() == "Compensações":
            in_offsets = True
            continue

        if in_offsets:
            offset_match = re.search(
                r"Número do Processo\s+(.+?)\s+Tipo\s+(.+?)\s+Valor\s+(.+)$",
                line,
            )
            if offset_match:
                offsets.append(
                    Compensacao(
                        codigo_receita=current.codigo_receita,
                        descricao_receita=current.descricao,
                        periodo_apuracao=current.periodo_apuracao,
                        numero_processo=offset_match.group(1).strip(),
                        tipo=offset_match.group(2).strip(),
                        valor=offset_match.group(3).strip(),
                    )
                )
                continue

        if any(key in line for key in ["Débito", "Crédito", "Compensação", "Saldo", "Deduções"]):
            current.outros_campos.append(line)

    if current and current.codigo_receita:
        blocks.append(current)

    return blocks, offsets


def extract_all(pdf_path: Path) -> Dict[str, object]:
    lines = extract_pdf_lines(pdf_path)
    header = extract_header(lines)
    blocks, offsets = extract_debits_credits_and_offsets(lines)

    return {
        "header": header,
        "blocks": blocks,
        "offsets": offsets,
    }


def _debits_rows(header: Dict[str, str], blocks: List[BlocoDebitoCredito]) -> List[List[str]]:
    rows = [
        ["Nome do Contribuinte", header.get("nome_contribuinte", "")],
        ["CNPJ", header.get("cnpj", "")],
        [],
        [
            "Código Receita",
            "Descrição",
            "Período Apuração",
            "CNPJ Débito",
            "Município Débito",
            "Débito Apurado",
            "Créditos Compensação",
            "Saldo a Pagar",
            "Deduções",
            "Outros Campos",
        ],
    ]

    for block in blocks:
        rows.append(
            [
                block.codigo_receita,
                block.descricao,
                block.periodo_apuracao,
                block.cnpj_debito,
                block.municipio_debito,
                block.debito_apurado,
                block.creditos_compensacao,
                block.saldo_a_pagar,
                " | ".join(block.deducoes),
                " | ".join(block.outros_campos),
            ]
        )

    return rows


def _offset_rows(header: Dict[str, str], offsets: List[Compensacao]) -> List[List[str]]:
    rows = [
        ["Nome do Contribuinte", header.get("nome_contribuinte", "")],
        ["CNPJ", header.get("cnpj", "")],
        [],
        [
            "Código Receita",
            "Descrição Receita",
            "Período Apuração",
            "Número do Processo",
            "Tipo",
            "Valor",
        ],
    ]

    for item in offsets:
        rows.append(
            [
                item.codigo_receita,
                item.descricao_receita,
                item.periodo_apuracao,
                item.numero_processo,
                item.tipo,
                item.valor,
            ]
        )

    return rows




def _to_excel_number(raw_value: str):
    value = (raw_value or "").strip()
    if not value:
        return None

    normalized = value.replace(".", "").replace(",", ".")
    if not re.fullmatch(r"-?\d+(?:\.\d+)?", normalized):
        return None

    try:
        return float(normalized)
    except ValueError:
        return None


def _to_excel_period(raw_period: str):
    period = (raw_period or "").strip()
    if not period:
        return ""

    if re.fullmatch(r"\d{2}/\d{4}", period):
        month, year = period.split("/")
        try:
            return datetime(int(year), int(month), 1)
        except ValueError:
            return period

    return period


def _apply_excel_types_debits(ws):
    for row_idx in range(5, ws.max_row + 1):
        period_cell = ws.cell(row=row_idx, column=3)
        if isinstance(period_cell.value, str):
            converted_period = _to_excel_period(period_cell.value)
            period_cell.value = converted_period
            if isinstance(converted_period, datetime):
                period_cell.number_format = "mm/yyyy"

        for col_idx in (6, 7, 8):
            num_cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(num_cell.value, str):
                converted_number = _to_excel_number(num_cell.value)
                if converted_number is not None:
                    num_cell.value = converted_number
                    num_cell.number_format = "#,##0.00"


def _apply_excel_types_offsets(ws):
    for row_idx in range(5, ws.max_row + 1):
        period_cell = ws.cell(row=row_idx, column=3)
        if isinstance(period_cell.value, str):
            converted_period = _to_excel_period(period_cell.value)
            period_cell.value = converted_period
            if isinstance(converted_period, datetime):
                period_cell.number_format = "mm/yyyy"

        value_cell = ws.cell(row=row_idx, column=6)
        if isinstance(value_cell.value, str):
            converted_number = _to_excel_number(value_cell.value)
            if converted_number is not None:
                value_cell.value = converted_number
                value_cell.number_format = "#,##0.00"

def export_result(header: Dict[str, str], blocks: List[BlocoDebitoCredito], offsets: List[Compensacao], output_path: Path) -> Dict[str, object]:
    debits = _debits_rows(header, blocks)
    offs = _offset_rows(header, offsets)

    openpyxl_available = importlib.util.find_spec("openpyxl") is not None

    if openpyxl_available and output_path.suffix.lower() == ".xlsx":
        from openpyxl import Workbook

        workbook = Workbook()
        ws_deb = workbook.active
        ws_deb.title = "Debitos_Creditos"
        for row in debits:
            ws_deb.append(row)
        _apply_excel_types_debits(ws_deb)

        ws_off = workbook.create_sheet("Compensacoes")
        for row in offs:
            ws_off.append(row)
        _apply_excel_types_offsets(ws_off)

        workbook.save(output_path)
        return {
            "format": "xlsx",
            "main_path": str(output_path),
            "extra_paths": [],
        }

    base = output_path.with_suffix("")
    debit_path = Path(f"{base}_debitos_creditos.csv")
    offset_path = Path(f"{base}_compensacoes.csv")

    with debit_path.open("w", newline="", encoding="utf-8") as file_deb:
        csv.writer(file_deb, delimiter=";").writerows(debits)

    with offset_path.open("w", newline="", encoding="utf-8") as file_off:
        csv.writer(file_off, delimiter=";").writerows(offs)

    return {
        "format": "csv",
        "main_path": str(debit_path),
        "extra_paths": [str(offset_path)],
    }
