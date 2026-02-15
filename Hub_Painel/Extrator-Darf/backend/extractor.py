import re
from datetime import datetime

import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import numbers


def valor_str_para_float(valor_str):
    """Converte string tipo '1.234,56' ou '-' para float (ex: 1234.56)."""
    if not valor_str or valor_str.strip() == "-" or valor_str.strip() == "":
        return 0.0
    valor_limpo = valor_str.replace(".", "").replace(",", ".")
    try:
        return float(valor_limpo)
    except ValueError:
        return 0.0


def extrair_dados(caminho_pdf, codigo_alvo=""):
    doc = fitz.open(caminho_pdf)
    info = {}
    dados = []

    for pag in doc:
        blocos = [
            " ".join(
                span["text"] for linha in b["lines"] for span in linha["spans"]
            )
            for b in pag.get_text("dict")["blocks"]
            if "lines" in b
        ]

        texto_pagina = "\n".join(blocos)

        # Extrair CNPJ e Razao Social uma vez só
        if "CNPJ" not in info:
            m = re.search(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}", texto_pagina)
            if m:
                info["CNPJ"] = m.group()
                linha_cnpj = next((l for l in blocos if info["CNPJ"] in l), "")
                info["Razao Social"] = linha_cnpj.split(info["CNPJ"])[-1].strip()

        # Extrair data de apuração da página a partir do bloco que contenha "Período Apuração"
        data_apuracao_pagina = ""
        for b in blocos:
            if "Período Apuração" in b:
                m = re.search(r"(\d{2}/\d{2}/\d{4})", b)
                if m:
                    data_apuracao_pagina = m.group(1)
                    break

        # Fallback: primeira data da página, se não encontrou data de apuração específica
        if not data_apuracao_pagina:
            m = re.search(r"\b(\d{2}/\d{2}/\d{4})\b", texto_pagina)
            if m:
                data_apuracao_pagina = m.group(1)

        i = 0
        while i < len(blocos):
            bloco = blocos[i]

            m_codigo = re.match(r"(\d{4})\s", bloco)
            if not m_codigo:
                i += 1
                continue

            codigo = m_codigo.group(1)
            if codigo_alvo and codigo != codigo_alvo:
                i += 1
                continue

            valores = re.findall(r"\d{1,3}(?:\.\d{3})*,\d{2}|-", bloco)
            if len(valores) < 4:
                i += 1
                continue

            descricao_final = None
            j = i + 1
            while j < len(blocos):
                prox = blocos[j]
                if re.match(r"\d{4}\s", prox) or re.search(
                    r"\d{1,3}(?:\.\d{3})*,\d{2}", prox
                ):
                    break
                if re.match(r"\d+\s*-\s*", prox):
                    descricao_final = prox.strip()
                    break
                j += 1

            if not descricao_final:
                i = j
                continue

            linha = [
                data_apuracao_pagina,
                codigo,
                descricao_final,
                valores[-4],  # Principal
                valores[-3],  # Multa
                valores[-2],  # Juros
                valores[-1],  # Total
            ]

            dados.append(linha)
            i = j

    doc.close()
    return info, dados


def salvar_em_excel(info, linhas, nome_arquivo="resultado.xlsx"):
    wb = Workbook()
    ws = wb.active

    # Cabeçalho empresa
    ws["A1"] = "EMPRESA"
    ws["A2"] = info.get("CNPJ", "")
    ws["A3"] = info.get("Razao Social", "")

    # Cabeçalho da tabela em B4
    headers = [
        "Período Apuração",
        "Código",
        "Descrição",
        "Principal",
        "Multas",
        "Juros",
        "Total",
    ]
    for col, h in enumerate(headers, start=2):
        ws.cell(row=4, column=col, value=h)

    for i, linha in enumerate(linhas, start=5):
        for j, item in enumerate(linha, start=2):
            cell = ws.cell(row=i, column=j)

            # Coluna 2 = Período Apuração  ➜  converte p/ data
            if j == 2 and item:
                try:
                    cell.value = datetime.strptime(item, "%d/%m/%Y").date()
                    cell.number_format = "DD/MM/YYYY"
                except ValueError:
                    cell.value = item

            # Colunas 5 a 8 (Principal, Multas, Juros, Total) ➜ converter para float
            elif j in [5, 6, 7, 8]:
                fval = valor_str_para_float(item)
                cell.value = fval
                cell.number_format = numbers.FORMAT_NUMBER_00  # 2 casas decimais

            else:
                cell.value = item

    wb.save(nome_arquivo)
    return nome_arquivo
